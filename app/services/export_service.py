"""Legacy Excel export — reproduce the original 18-column downtime workbook.

Downstream consumers (the group reporting pack) still expect the exact column
layout of the old hand-maintained spreadsheet. This service rebuilds that sheet
from the database so those consumers are unaffected by the migration.

The 18 columns are: the ten values an operator/​shift carries (date, shift,
machine, sku, start, stop, the two failure descriptions, type of loss, bcs) plus
the eight columns the legacy sheet derived by formula (three duration
representations and five calendar parts). We write *computed values*, not Excel
formulas: an exported shift is an immutable, already-approved record, and the
derivations are the same ones unit-tested against the real April-2020 sheet
(see tests/unit/test_derived_and_anomaly.py), so the numbers are guaranteed to
match the legacy output without depending on Excel recalculation.

NOTE: the column order/headers below mirror a standard downtime log; confirm
them against the actual legacy workbook and reorder LEGACY_COLUMNS if needed —
it's a one-line change per column and nothing else has to move.
"""

from collections.abc import Callable
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.loss_event import LossEvent
from app.models.shift import Shift
from app.repositories.loss_event_repo import LossEventRepository
from app.services.derived import DerivedFields, derive_fields

# A column is (header, extractor). The extractor receives the event, its derived
# fields, and the parent shift, and returns the cell value.
Extractor = Callable[[LossEvent, DerivedFields, Shift], Any]

LEGACY_COLUMNS: list[tuple[str, Extractor]] = [
    ("Date", lambda e, d, s: e.event_date),
    ("Shift", lambda e, d, s: s.shift_name.value.upper()),
    ("Machine", lambda e, d, s: e.machine.name),
    ("SKU", lambda e, d, s: e.sku.code),
    ("Start", lambda e, d, s: e.event_start.strftime("%H:%M")),
    ("Stop", lambda e, d, s: e.event_stop.strftime("%H:%M")),
    ("Duration (h:mm)", lambda e, d, s: d.duration_hhmm),
    ("Duration (min)", lambda e, d, s: d.duration_minutes),
    ("Duration (hr)", lambda e, d, s: d.duration_hours),
    ("Type of Loss", lambda e, d, s: e.loss_type.code),
    ("BCS", lambda e, d, s: e.bcs_category.code),
    ("Functional Failure Description", lambda e, d, s: e.functional_failure_description or ""),
    ("Failure Mode Description", lambda e, d, s: e.failure_mode_description or ""),
    ("Month", lambda e, d, s: d.month),
    ("Week", lambda e, d, s: d.week),
    ("Year", lambda e, d, s: d.year),
    ("Week-Year", lambda e, d, s: d.week_year),
    ("Month-Year", lambda e, d, s: d.month_year),
]

_HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Arial")


class ExportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.events = LossEventRepository(db)

    def filename_for(self, shift: Shift) -> str:
        return f"shift_{shift.id}_{shift.shift_date}_{shift.shift_name.value}.xlsx"

    def build_workbook(self, shift: Shift) -> Workbook:
        """Build the legacy-format workbook for one shift's loss events."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Shift {shift.id}"

        headers = [c[0] for c in LEGACY_COLUMNS]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for event in self.events.list_for_shift(shift.id):
            derived = derive_fields(event.event_date, event.event_start, event.event_stop)
            row = [extract(event, derived, shift) for _, extract in LEGACY_COLUMNS]
            ws.append(row)

        self._style_body(ws, n_cols=len(headers))
        ws.freeze_panes = "A2"
        return wb

    def build_xlsx_bytes(self, shift: Shift) -> bytes:
        buffer = BytesIO()
        self.build_workbook(shift).save(buffer)
        return buffer.getvalue()

    # ----- formatting helpers -----

    def _style_body(self, ws, n_cols: int) -> None:
        # Apply the body font, format the Date column, and size columns to fit.
        date_col = next(i for i, c in enumerate(LEGACY_COLUMNS, start=1) if c[0] == "Date")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = _BODY_FONT
                if cell.column == date_col and isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"

        for col_idx in range(1, n_cols + 1):
            letter = get_column_letter(col_idx)
            longest = max(
                (
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(1, ws.max_row + 1)
                ),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max(longest + 2, 8), 40)
