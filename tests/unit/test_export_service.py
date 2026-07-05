"""Unit tests for the legacy 18-column Excel export.

The event used here matches the historical 1-Apr-2020 sheet, so these tests also
prove the exported derived columns reproduce the legacy values exactly.
"""

from datetime import date, time
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.database import Base, SessionLocal, engine
from app.models.lookup import SKU, BCSCategory, LossType, Machine
from app.models.loss_event import LossEvent
from app.models.shift import Role, Shift, ShiftName, ShiftStatus, User
from app.services.export_service import LEGACY_COLUMNS, ExportService

EXPECTED_HEADERS = [
    "Date",
    "Shift",
    "Machine",
    "SKU",
    "Start",
    "Stop",
    "Duration (h:mm)",
    "Duration (min)",
    "Duration (hr)",
    "Type of Loss",
    "BCS",
    "Functional Failure Description",
    "Failure Mode Description",
    "Month",
    "Week",
    "Year",
    "Week-Year",
    "Month-Year",
]


@pytest.fixture(autouse=True)
def fresh_db():
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    yield


def _seed_shift_with_event(db):
    op = User(username="op", password_hash="x", role=Role.OPERATOR)
    db.add(op)
    db.flush()
    sku = SKU(code="Brew Lager(33cl)")
    machine = Machine(name="FillerCrowner")
    bcs = BCSCategory(code="Breakdown")
    lt = LossType(code="Breakdown", suggested_bcs_code="Breakdown")
    db.add_all([sku, machine, bcs, lt])
    db.flush()
    shift = Shift(
        shift_date=date(2020, 4, 1),
        shift_name=ShiftName.A,
        line="Line 1",
        operator_id=op.id,
        status=ShiftStatus.APPROVED,
    )
    db.add(shift)
    db.flush()
    db.add(
        LossEvent(
            shift_id=shift.id,
            event_date=date(2020, 4, 1),
            event_start=time(7, 0),
            event_stop=time(8, 0),
            sku_id=sku.id,
            machine_id=machine.id,
            bcs_category_id=bcs.id,
            loss_type_id=lt.id,
            functional_failure_description="Lift cylinder dropped",
            failure_mode_description="Loosened holding bolt",
            created_by_id=op.id,
        )
    )
    db.commit()
    return shift


def test_has_exactly_eighteen_columns():
    assert len(LEGACY_COLUMNS) == 18
    assert [c[0] for c in LEGACY_COLUMNS] == EXPECTED_HEADERS


def test_header_row_matches_legacy():
    db = SessionLocal()
    shift = _seed_shift_with_event(db)
    wb = ExportService(db).build_workbook(shift)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, 19)]
    assert headers == EXPECTED_HEADERS
    db.close()


def test_event_row_values_match_legacy():
    db = SessionLocal()
    shift = _seed_shift_with_event(db)
    wb = ExportService(db).build_workbook(shift)
    ws = wb.active
    row = {ws.cell(row=1, column=i).value: ws.cell(row=2, column=i).value for i in range(1, 19)}

    assert row["Shift"] == "A"
    assert row["Machine"] == "FillerCrowner"
    assert row["SKU"] == "Brew Lager(33cl)"
    assert row["Start"] == "07:00"
    assert row["Stop"] == "08:00"
    assert row["Duration (h:mm)"] == "1:00"
    assert row["Duration (min)"] == 60
    assert row["Duration (hr)"] == 1.0
    assert row["Type of Loss"] == "Breakdown"
    assert row["BCS"] == "Breakdown"
    assert row["Functional Failure Description"] == "Lift cylinder dropped"
    # Derived calendar parts, verified against the legacy sheet.
    assert row["Month"] == 4
    assert row["Week"] == 14
    assert row["Year"] == 2020
    assert row["Week-Year"] == "Wk14-2020"
    assert row["Month-Year"] == "M4-2020"
    db.close()


def test_empty_shift_exports_header_only():
    db = SessionLocal()
    op = User(username="op", password_hash="x", role=Role.OPERATOR)
    db.add(op)
    db.flush()
    shift = Shift(
        shift_date=date(2026, 6, 1),
        shift_name=ShiftName.B,
        operator_id=op.id,
        status=ShiftStatus.PENDING_REVIEW,
    )
    db.add(shift)
    db.commit()
    wb = ExportService(db).build_workbook(shift)
    ws = wb.active
    assert ws.max_row == 1  # header only
    db.close()


def test_build_xlsx_bytes_is_loadable():
    db = SessionLocal()
    shift = _seed_shift_with_event(db)
    content = ExportService(db).build_xlsx_bytes(shift)
    assert content[:2] == b"PK"  # xlsx is a zip
    wb = load_workbook(BytesIO(content))
    assert wb.active.max_row == 2
    db.close()


def test_filename_includes_shift_metadata():
    db = SessionLocal()
    shift = _seed_shift_with_event(db)
    name = ExportService(db).filename_for(shift)
    assert name.endswith(".xlsx")
    assert "2020-04-01" in name
    db.close()
