"""Integration tests for B-03: the legacy Excel export route."""

from datetime import date, time
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import Base, SessionLocal, engine
from app.main import app
from app.models.lookup import SKU, BCSCategory, LossType, Machine
from app.models.loss_event import LossEvent
from app.models.shift import Role, Shift, ShiftName, ShiftStatus, User
from app.services.auth_service import hash_password

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(autouse=True)
def setup_db():
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    db = SessionLocal()
    op = User(username="operator1", password_hash=hash_password("operator1"), role=Role.OPERATOR)
    sup = User(username="super1", password_hash=hash_password("super1"), role=Role.SUPERVISOR)
    db.add_all([op, sup])
    db.flush()
    sku = SKU(code="Brew Lager(33cl)")
    machine = Machine(name="FillerCrowner")
    bcs = BCSCategory(code="Breakdown")
    lt = LossType(code="Breakdown", suggested_bcs_code="Breakdown")
    db.add_all([sku, machine, bcs, lt])
    db.flush()
    shift = Shift(
        shift_date=date(2026, 6, 21),
        shift_name=ShiftName.B,
        line="Line 1",
        operator_id=op.id,
        status=ShiftStatus.APPROVED,
    )
    db.add(shift)
    db.flush()
    db.add(
        LossEvent(
            shift_id=shift.id,
            event_date=date(2026, 6, 21),
            event_start=time(7, 0),
            event_stop=time(7, 45),
            sku_id=sku.id,
            machine_id=machine.id,
            bcs_category_id=bcs.id,
            loss_type_id=lt.id,
            created_by_id=op.id,
        )
    )
    db.commit()
    sid = shift.id
    db.close()
    yield sid


def _client() -> TestClient:
    return TestClient(app)


def _login(c, username):
    c.post("/login", data={"username": username, "password": username})


def test_export_requires_login(setup_db):
    c = _client()
    r = c.get(f"/supervisor/shift/{setup_db}/export.xlsx", follow_redirects=False)
    assert r.status_code == 303
    assert r.headers["location"].startswith("/login")


def test_operator_forbidden_from_export(setup_db):
    c = _client()
    _login(c, "operator1")
    r = c.get(f"/supervisor/shift/{setup_db}/export.xlsx")
    assert r.status_code == 403


def test_export_returns_xlsx_with_attachment_header(setup_db):
    c = _client()
    _login(c, "super1")
    r = c.get(f"/supervisor/shift/{setup_db}/export.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "attachment" in r.headers["content-disposition"]
    assert ".xlsx" in r.headers["content-disposition"]


def test_exported_workbook_has_18_columns_and_event_row(setup_db):
    c = _client()
    _login(c, "super1")
    r = c.get(f"/supervisor/shift/{setup_db}/export.xlsx")
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert len(headers) == 18
    assert headers[0] == "S/N."
    assert headers[1] == "Date"
    assert ws.max_row == 2  # header + one event
    # spot-check the 45-minute duration landed
    row = {ws.cell(row=1, column=i).value: ws.cell(row=2, column=i).value for i in range(1, 19)}
    assert row["Duration (min)"] == 45
    assert row["Machine"] == "FillerCrowner"


def test_export_missing_shift_404(setup_db):
    c = _client()
    _login(c, "super1")
    r = c.get("/supervisor/shift/9999/export.xlsx")
    assert r.status_code == 404
